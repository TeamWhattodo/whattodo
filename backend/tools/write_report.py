import os
import json
import re
from datetime import datetime
from pathlib import Path
from backend.agents.llm_client import complete

from openpyxl import load_workbook


try:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

GOV_REPORT_SYSTEM = """
당신은 한국 공공기관/대기업 기안서(보고서) 작성 전문가입니다.
주어진 데이터를 바탕으로 공문서 스타일의 보고서를 작성하세요.
반드시 아래 JSON 스키마를 엄격하게 지켜서 출력하세요. 다른 설명 텍스트는 절대 포함하지 마세요.

{
  "title": "보고서 대제목 (예: 코로나 재택근무 대비를 위한 플레이스테이션 5 구매계획)",
  "meta": "작성일 및 작성자 (예: 2024. 9. 4.(수) 기획처 사업부 홍길동)",
  "sections": [
    {
      "heading": "1. 대제목 (예: 1. 개요, 2. 배경 및 사전검토)",
      "items": [
        {
          "type": "square",  
          "text": "본문 내용",
          "highlight": "none" 
        }
      ]
    }
  ]
}

* 규칙 *
1. type은 "square"(□), "circle"(○), "hyphen"(-) 중 하나. 대분류는 square, 중분류는 circle, 소분류는 hyphen.
2. highlight는 "none", "blue", "red", "green" 중 하나. 중요한 문구 전체를 강조할 때만 색상 지정.
"""

DAILY_REPORT_SYSTEM = """
당신은 '일일 업무보고서' 작성 전문가입니다.
주어진 데이터를 바탕으로 일일 업무보고서 항목을 작성하세요.
반드시 아래 JSON 스키마를 엄격하게 지켜서 출력하세요. 다른 텍스트는 절대 포함하지 마세요.

{
  "date": "2024. 9. 4",
  "author": "작성자 이름",
  "department": "소속팀명",
  "position": "직급 (예: 사원, 대리)",
  "today_tasks": "오늘의 주요업무 내용 (텍스트로 길게 작성 가능)",
  "pending_tasks": "미종결 업무사항 내용",
  "tomorrow_tasks": "익일 업무계획 내용",
  "other_notes": "기타사항 내용"
}
"""

WEEKLY_REPORT_SYSTEM = """
당신은 '주간 업무보고서' 작성 전문가입니다.
주어진 데이터를 바탕으로 주간 업무보고서 항목을 작성하세요.
반드시 아래 JSON 스키마를 엄격하게 지켜서 출력하세요. 다른 텍스트는 절대 포함하지 마세요.

* 작성 가이드 *
- 모든 내용은 배열(리스트) 형태로 항목별로 분리해서 작성한다.
- 각 항목은 간결하게 핵심만 한 줄로 작성한다. 서술형 금지.
- 항목이 없으면 빈 배열([])로 작성한다.

{
  "date": "2024년 9월 4일",
  "department": "기획팀",
  "position": "대리",
  "author": "홍길동",
  "past_week_tasks": ["전주 추진사항 항목1", "항목2"],
  "next_week_plans": ["차주 계획사항 항목1", "항목2"],
  "special_notes": ["특이사항 항목1"],
  "instructions": ["지시사항 항목1"]
}
"""

MONTHLY_REPORT_SYSTEM = """
당신은 '월간 업무보고서' 작성 전문가입니다.
주어진 데이터를 바탕으로 월간 업무보고서 항목을 작성하세요.
반드시 아래 JSON 스키마를 엄격하게 지켜서 출력하세요. 다른 텍스트는 절대 포함하지 마세요.

* 작성 가이드 *
- '금월 진행업무'와 '차월 업무계획'은 여러 건의 업무일 경우 배열(리스트) 형태로 각각 분리해서 작성합니다.
- 각 업무의 제목, 상태, 결과, 비고 등을 짧고 명확하게 작성합니다.

{
  "department": "기획팀",
  "author": "홍길동",
  "date": "2024년 9월",
  "this_month_tasks": [
    {"task": "금월 업무내용 1", "status": "완료", "result": "결과 요약", "note": "-"}
  ],
  "next_month_plans": [
    {"task": "차월 업무내용 1", "date": "예정일자", "issue": "예상 문제점", "note": "-"}
  ],
  "opinions": "기타 의견 텍스트"
}
"""

_EXPENSE_OUTPUT_DIR    = "outputs"
_EXPENSE_TEMPLATE_PATH = str(Path(__file__).resolve().parents[2] / "backend/db/data/Form/경비정산서 양식.xlsx")
_EXPENSE_CATEGORY_MAP  = {
    "식비":   "식비",
    "숙박비": "숙박비",
    "유류비": "교통비",
    "출장비": "기타",
    "기타":   "기타",
}


def _write_expense_xlsx(
    items: list[dict],
    total: int | float,
    report_id: str,
    author: str = "",
    department: str = "",
) -> str:
    path = f"{_EXPENSE_OUTPUT_DIR}/{report_id}.xlsx"
    wb = load_workbook(_EXPENSE_TEMPLATE_PATH)
    ws = wb.active
    if ws is None:
        raise ValueError(f"활성 시트 없음: {_EXPENSE_TEMPLATE_PATH}")

    data_start_row = None
    total_row      = None
    dept_row       = None
    name_row       = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "내역":
                data_start_row = cell.row
            elif cell.value == "경비 총액":
                total_row = cell.row
            elif cell.value == "소속":
                dept_row = cell.row
            elif cell.value == "성명":
                name_row = cell.row

    if data_start_row is None or total_row is None:
        raise ValueError("템플릿에서 '내역' 또는 '경비 총액' 행을 찾지 못함")

    if dept_row and department:
        ws[f"D{dept_row}"] = department
    if name_row and author:
        ws[f"I{name_row}"] = author

    for i, item in enumerate(items):
        row = data_start_row + i
        ws[f"C{row}"] = item["date"]
        ws[f"E{row}"] = _EXPENSE_CATEGORY_MAP.get(item["category"], "기타")
        ws[f"G{row}"] = item["amount"]
        memo = item.get("memo")
        ws[f"I{row}"] = "" if not memo or memo == "null" else memo

    ws[f"D{total_row}"] = total
    wb.save(path)
    return path

def write_report(report_type: str, data: dict | list) -> dict:
    """
    report_type: "expense_report" | "briefing" | "daily_summary" | "kpi_weekly" | "monthly_summary" | "billing"
    data: 보고서에 포함할 항목들
    반환: {"report_type": ..., "content": str, "pdf_path"/"xlsx_path": str}
    """
    if report_type == "expense_report":
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                data = {}
        items      = data.get("items", []) if isinstance(data, dict) else data
        total      = sum(i.get("amount", 0) for i in items if isinstance(i.get("amount"), (int, float)))
        author     = data.get("author", "") if isinstance(data, dict) else ""
        department = data.get("department", "") if isinstance(data, dict) else ""
        os.makedirs(_EXPENSE_OUTPUT_DIR, exist_ok=True)
        report_id = f"경비정산서_{datetime.now().strftime('%Y%m%d')}"
        xlsx_path = _write_expense_xlsx(items, total, report_id, author=author, department=department)
        return {
            "report_type":  "expense_report",
            "total_amount": total,
            "items":        items,
            "xlsx_path":    xlsx_path,
        }

    if isinstance(data, str):
        data_str = data
    else:
        try:
            data_str = json.dumps(data, ensure_ascii=False, default=str)
        except Exception:
            data_str = str(data)

    is_daily = (report_type == "daily_summary")
    is_weekly = (report_type == "kpi_weekly")
    is_monthly = (report_type == "monthly_summary")

    if is_monthly:
        system = MONTHLY_REPORT_SYSTEM
    elif is_weekly:
        system = WEEKLY_REPORT_SYSTEM
    elif is_daily:
        system = DAILY_REPORT_SYSTEM
    else:
        system = GOV_REPORT_SYSTEM
        
    prompt = f"[{report_type}] 데이터를 바탕으로 JSON 보고서를 작성해줘:\n{data_str}"
    
    try:
        raw_output = complete(prompt, tier="smart", system=system)
        
        start_idx = raw_output.find("{")
        end_idx = raw_output.rfind("}")
        if start_idx != -1 and end_idx != -1:
            raw_output = raw_output[start_idx:end_idx+1]
        
        parsed_data = json.loads(raw_output)
        
        if is_monthly:
            md_content = f"# 월간 업무보고서\n"
            md_content += f"**부서:** {parsed_data.get('department', '')} | **작성자:** {parsed_data.get('author', '')} | **일자:** {parsed_data.get('date', '')}\n\n"
            md_content += f"### ■ 금월 진행업무\n"
            for t in parsed_data.get('this_month_tasks', []):
                md_content += f"- **{t.get('task', '')}** (상태: {t.get('status', '')}, 결과: {t.get('result', '')}, 비고: {t.get('note', '')})\n"
            md_content += f"\n### ■ 차월 업무계획\n"
            for t in parsed_data.get('next_month_plans', []):
                md_content += f"- **{t.get('task', '')}** (예정일자: {t.get('date', '')}, 이슈: {t.get('issue', '')}, 비고: {t.get('note', '')})\n"
            md_content += f"\n### ■ 기타 의견\n{parsed_data.get('opinions', '')}"
        elif is_weekly:
            def _md_list(val) -> str:
                if isinstance(val, list):
                    return "\n".join(f"- {item}" for item in val) if val else "-"
                return str(val) if val else "-"
            md_content = f"# 주간 업무보고서\n"
            md_content += f"**부서:** {parsed_data.get('department', '')} | **직책:** {parsed_data.get('position', '')} | **작성자:** {parsed_data.get('author', '')} | **작성일:** {parsed_data.get('date', '')}\n\n"
            md_content += f"### ▶ 전주 추진사항\n{_md_list(parsed_data.get('past_week_tasks', []))}\n\n"
            md_content += f"### ▶ 차주 계획사항\n{_md_list(parsed_data.get('next_week_plans', []))}\n\n"
            md_content += f"### ▶ 특이사항\n{_md_list(parsed_data.get('special_notes', []))}\n\n"
            md_content += f"### ▶ 지시사항\n{_md_list(parsed_data.get('instructions', []))}"
        elif is_daily:
            md_content = f"# 일일 업무보고서\n"
            md_content += f"**작성자:** {parsed_data.get('department', '')} {parsed_data.get('position', '')} {parsed_data.get('author', '')} ({parsed_data.get('date', '')})\n\n"
            md_content += f"### ▶ 오늘의 주요업무\n{parsed_data.get('today_tasks', '')}\n\n"
            md_content += f"### ▶ 미종결 업무사항\n{parsed_data.get('pending_tasks', '')}\n\n"
            md_content += f"### ▶ 익일 업무계획\n{parsed_data.get('tomorrow_tasks', '')}\n\n"
            md_content += f"### ▶ 기타사항\n{parsed_data.get('other_notes', '')}"
        else:
            md_content = f"# {parsed_data.get('title', '보고서')}\n"
            md_content += f"*{parsed_data.get('meta', '')}*\n\n"
            for sec in parsed_data.get('sections', []):
                md_content += f"### {sec.get('heading', '')}\n"
                for item in sec.get('items', []):
                    bullet = "□" if item.get('type') == 'square' else "○" if item.get('type') == 'circle' else "-"
                    text = item.get('text', '')
                    md_content += f"{bullet} {text}\n"
                md_content += "\n"
            
        result = {"report_type": report_type, "content": md_content}
        
        if REPORTLAB_AVAILABLE:
            if is_monthly:
                pdf_path = _generate_monthly_report_pdf(parsed_data, report_type)
            elif is_weekly:
                pdf_path = _generate_weekly_report_pdf(parsed_data, report_type)
            elif is_daily:
                pdf_path = _generate_daily_report_pdf(parsed_data, report_type)
            else:
                pdf_path = _generate_gov_style_pdf(parsed_data, report_type)
                
            if pdf_path:
                result["pdf_path"] = pdf_path
                
        return result
    except Exception as e:
        return {"report_type": report_type, "content": f"보고서 생성 실패: {str(e)}", "error": str(e)}

_FONT_CANDIDATES = [
    ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
     "NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
    ("Malgun",      "C:/Windows/Fonts/malgun.ttf",
     "Malgun-Bold", "C:/Windows/Fonts/malgunbd.ttf"),
    ("AppleGothic", "/Library/Fonts/AppleGothic.ttf",
     "AppleGothic", "/Library/Fonts/AppleGothic.ttf"),
    ("AppleGothic", "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
     "AppleGothic", "/System/Library/Fonts/Supplemental/AppleGothic.ttf"),
]

def _setup_fonts() -> tuple[str, str]:
    for reg_name, reg_path, bold_name, bold_path in _FONT_CANDIDATES:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(reg_name, reg_path))
                if bold_name != reg_name and os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                return reg_name, bold_name
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"

def _generate_gov_style_pdf(data: dict, report_type: str) -> str:
    try:
        os.makedirs("outputs", exist_ok=True)
        filename = f"기안서_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join("outputs", filename)
        font_normal, font_bold = _setup_fonts()
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2.5*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(name="GovTitle", parent=styles["Normal"], fontName=font_bold, fontSize=18, leading=22, alignment=1)
        meta_style = ParagraphStyle(name="GovMeta", parent=styles["Normal"], fontName=font_normal, fontSize=10, leading=14, alignment=2)
        heading_style = ParagraphStyle(name="GovHeading", parent=styles["Normal"], fontName=font_bold, fontSize=13, leading=20, spaceBefore=12, spaceAfter=6)
        def get_item_style(indent_level=0):
            return ParagraphStyle(name=f"GovItem_{indent_level}", parent=styles["Normal"], fontName=font_normal, fontSize=11, leading=16, leftIndent=indent_level, wordWrap='CJK')
            
        story = []
        story.append(Paragraph(data.get("title", "보고서"), title_style))
        story.append(Spacer(1, 0.3*cm))
        line_data = [[Paragraph(data.get("meta", ""), meta_style)]]
        line_table = Table(line_data, colWidths=[doc.width])
        line_table.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,-1), 2, colors.black),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.black),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(line_table)
        story.append(Spacer(1, 0.5*cm))
        
        for sec in data.get("sections", []):
            story.append(Paragraph(sec.get("heading", ""), heading_style))
            for item in sec.get("items", []):
                i_type = item.get("type", "square")
                i_text = item.get("text", "")
                highlight = item.get("highlight", "none")
                color_hex = "#000000"
                if highlight == "blue": color_hex = "#0000FF"
                elif highlight == "red": color_hex = "#FF0000"
                elif highlight == "green": color_hex = "#008000"
                formatted_text = f"<font color='{color_hex}'>{i_text}</font>"
                bullet = "□"
                indent = 10
                if i_type == "circle": bullet = "○"; indent = 25
                elif i_type == "hyphen": bullet = "-"; indent = 40
                story.append(Paragraph(f"{bullet} {formatted_text}", get_item_style(indent)))
                story.append(Spacer(1, 0.1*cm))
        doc.build(story)
        return pdf_path
    except Exception as e:
        print(f"PDF 생성 실패: {e}")
        return ""

def _generate_daily_report_pdf(data: dict, report_type: str) -> str:
    try:
        os.makedirs("outputs", exist_ok=True)
        filename = f"일일보고서_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join("outputs", filename)
        font_normal, font_bold = _setup_fonts()
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        title_style = ParagraphStyle(name="DTitle", fontName=font_bold, fontSize=18, alignment=0)
        table_normal = ParagraphStyle(name="TNormal", fontName=font_normal, fontSize=10, alignment=1)
        section_title = ParagraphStyle(name="SecTitle", fontName=font_bold, fontSize=11, alignment=0, textColor=colors.black)
        body_style = ParagraphStyle(name="DBody", fontName=font_normal, fontSize=10, leading=16, alignment=0)
        
        story = []
        cw = doc.width
        app_box_width = cw * 0.4
        t_title = Paragraph("일일 업무보고서", title_style)
        
        t_app_data = [[Paragraph("담당", table_normal), '', ''], ['', '', '']]
        t_app = Table(t_app_data, colWidths=[app_box_width/3.0]*3, rowHeights=[0.6*cm, 1.2*cm])
        t_app.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        
        t_header_data = [[t_title, t_app]]
        t_header = Table(t_header_data, colWidths=[cw*0.6, cw*0.4])
        t_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'BOTTOM')]))
        story.append(t_header)
        story.append(Spacer(1, 0.2*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceAfter=0.5*cm))
        
        lbl_w, val_w = cw * 0.15, cw * 0.35
        t_info_data = [
            [Paragraph("일 자", table_normal), Paragraph(data.get("date", ""), table_normal), Paragraph("작 성 자", table_normal), Paragraph(data.get("author", ""), table_normal)],
            [Paragraph("부 서", table_normal), Paragraph(data.get("department", ""), table_normal), Paragraph("직 급", table_normal), Paragraph(data.get("position", ""), table_normal)]
        ]
        t_info = Table(t_info_data, colWidths=[lbl_w, val_w, lbl_w, val_w], rowHeights=[0.8*cm, 0.8*cm])
        t_info.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('LINEABOVE', (0,0), (-1,0), 1, colors.black),
            ('LINEBELOW', (0,1), (-1,1), 1, colors.black),
            ('BACKGROUND', (0,0), (0,1), colors.whitesmoke),
            ('BACKGROUND', (2,0), (2,1), colors.whitesmoke),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        story.append(t_info)
        story.append(Spacer(1, 0.5*cm))
        
        def add_section(title, content):
            t_sec = Table([[Paragraph(title, section_title)]], colWidths=[cw], rowHeights=[0.8*cm])
            t_sec.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.whitesmoke), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LEFTPADDING', (0,0), (-1,-1), 5)]))
            story.append(t_sec)
            story.append(Spacer(1, 0.2*cm))
            safe_content = content.replace("\n", "<br/>")
            story.append(Paragraph(safe_content, body_style))
            story.append(Spacer(1, 2.5*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey, spaceAfter=0.5*cm))
            
        add_section("▶ 오늘의 주요업무", data.get("today_tasks", ""))
        add_section("▶ 미종결 업무사항", data.get("pending_tasks", ""))
        add_section("▶ 익일 업무계획", data.get("tomorrow_tasks", ""))
        add_section("▶ 기타사항", data.get("other_notes", ""))
        
        doc.build(story)
        return pdf_path
    except Exception as e:
        print(f"PDF 생성 실패: {e}")
        return ""

def _generate_weekly_report_pdf(data: dict, report_type: str) -> str:
    try:
        os.makedirs("outputs", exist_ok=True)
        filename = f"주간보고서_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join("outputs", filename)
        font_normal, font_bold = _setup_fonts()
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
        
        title_style = ParagraphStyle(name="WTitle", fontName=font_bold, fontSize=20, alignment=0)
        table_normal = ParagraphStyle(name="WTNormal", fontName=font_normal, fontSize=9, alignment=1)
        left_menu = ParagraphStyle(name="WMenu", fontName=font_bold, fontSize=9, alignment=1, textColor=colors.black, leading=14)
        col_header = ParagraphStyle(name="WColHeader", fontName=font_normal, fontSize=9, alignment=1, textColor=colors.black)
        body_style = ParagraphStyle(name="WBody", fontName=font_normal, fontSize=9, leading=16, alignment=0)
        
        story = []
        cw = doc.width
        app_box_width = cw * 0.4
        t_title = Paragraph("주간 업무보고서", title_style)
        
        t_app_data = [[Paragraph("담당", table_normal), '', ''], ['', '', '']]
        t_app = Table(t_app_data, colWidths=[app_box_width/3.0]*3, rowHeights=[0.6*cm, 1.2*cm])
        t_app.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        
        t_header = Table([[t_title, t_app]], colWidths=[cw*0.6, cw*0.4])
        t_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'BOTTOM')]))
        story.append(t_header)
        story.append(Spacer(1, 0.8*cm))
        
        info_html = f"작성일 : {data.get('date', '')} &nbsp;&nbsp;&nbsp;&nbsp; 부서 : {data.get('department', '')} &nbsp;&nbsp;&nbsp;&nbsp; 직책 : {data.get('position', '')} &nbsp;&nbsp;&nbsp;&nbsp; 작성자 : {data.get('author', '')}"
        p_info = Paragraph(info_html, ParagraphStyle(name="WInfo", fontName=font_normal, fontSize=9, alignment=0))
        story.append(p_info)
        story.append(Spacer(1, 0.2*cm))
        
        col1_w = cw * 0.1
        col2_w = cw * 0.9

        def _to_bullets(val) -> str:
            if isinstance(val, list):
                return "<br/>".join(f"• {item}" for item in val) if val else "-"
            return str(val).replace("\n", "<br/>") if val else "-"

        safe_past_tasks = _to_bullets(data.get("past_week_tasks", []))
        safe_next_plans = _to_bullets(data.get("next_week_plans", []))
        safe_spec       = _to_bullets(data.get("special_notes", []))
        safe_inst       = _to_bullets(data.get("instructions", []))

        t_main_data = [
            [Paragraph("전주<br/>실행<br/>사항", left_menu), Paragraph(safe_past_tasks, body_style), Spacer(1, 5.5*cm)],
            [Paragraph("차주<br/>계획", left_menu),          Paragraph(safe_next_plans, body_style), Spacer(1, 5.5*cm)],
            [Paragraph("특이<br/>사항", left_menu),          Paragraph(safe_spec, body_style),       Spacer(1, 3.5*cm)],
            [Paragraph("지시<br/>사항", left_menu),          Paragraph(safe_inst, body_style),       Spacer(1, 3.5*cm)],
        ]

        t_main = Table(t_main_data, colWidths=[col1_w, col2_w, 0], rowHeights=[None, None, None, None])
        t_main.setStyle(TableStyle([
            ('GRID', (0,0), (1,3), 0.5, colors.grey),
            ('LINEABOVE', (0,0), (1,0), 1, colors.black),
            ('BACKGROUND', (0,0), (0,3), colors.whitesmoke),
            ('VALIGN', (0,0), (0,3), 'MIDDLE'),
            ('ALIGN', (0,0), (0,3), 'CENTER'),
            ('VALIGN', (1,0), (1,3), 'TOP'),
            ('VALIGN', (1,4), (2,5), 'TOP'), 
            ('LEFTPADDING', (1,0), (1,3), 10),
            ('RIGHTPADDING', (1,0), (1,3), 10),
            ('TOPPADDING', (1,0), (1,3), 10),
            ('BOTTOMPADDING', (1,0), (1,3), 10),
            ('LEFTPADDING', (2,0), (2,-1), 0),
            ('RIGHTPADDING', (2,0), (2,-1), 0),
        ]))
        story.append(t_main)
        
        doc.build(story)
        return pdf_path
    except Exception as e:
        print(f"PDF 생성 실패: {e}")
        return ""

def _generate_monthly_report_pdf(data: dict, report_type: str) -> str:
    try:
        os.makedirs("outputs", exist_ok=True)
        filename = f"월간보고서_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join("outputs", filename)
        font_normal, font_bold = _setup_fonts()
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
        
        title_style = ParagraphStyle(name="MTitle", fontName=font_bold, fontSize=18, alignment=0)
        table_normal = ParagraphStyle(name="MTNormal", fontName=font_normal, fontSize=9, alignment=1)
        section_title = ParagraphStyle(name="MSecTitle", fontName=font_bold, fontSize=11, alignment=0)
        body_style = ParagraphStyle(name="MBody", fontName=font_normal, fontSize=9, leading=14, alignment=0)
        body_center = ParagraphStyle(name="MBodyCenter", fontName=font_normal, fontSize=9, leading=14, alignment=1)
        
        story = []
        cw = doc.width
        
        # Header
        app_box_width = cw * 0.45
        t_title = Paragraph("월간 업무보고서", title_style)
        
        t_app_data = [
            [Paragraph("부 서 명", table_normal), Paragraph(data.get("department", ""), table_normal)],
            [Paragraph("작 성 자", table_normal), Paragraph(data.get("author", ""), table_normal)],
            [Paragraph("일 자", table_normal), Paragraph(data.get("date", ""), table_normal)]
        ]
        t_app = Table(t_app_data, colWidths=[app_box_width*0.4, app_box_width*0.6], rowHeights=[0.8*cm]*3)
        t_app.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('LINEABOVE', (0,0), (-1,0), 1, colors.black),
            ('LINEBELOW', (0,-1), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        
        t_header = Table([[t_title, t_app]], colWidths=[cw*0.55, cw*0.45])
        t_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        story.append(t_header)
        story.append(Spacer(1, 0.5*cm))
        
        def make_section_table(items, col_headers, keys, min_rows=5):
            # Header Row
            row_data = [[Paragraph(h, table_normal) for h in col_headers] + ['']]
            
            # Body Rows
            for item in items:
                cells = []
                for idx, k in enumerate(keys):
                    val = item.get(k, "")
                    style = body_style if idx == 0 else body_center
                    cells.append(Paragraph(val, style))
                row_data.append(cells + [Spacer(1, 1.0*cm)])
                
            # Fill remaining rows
            while len(row_data) <= min_rows:
                row_data.append(['', '', '', '', Spacer(1, 1.0*cm)])
                
            col_w = [cw*0.45, cw*0.15, cw*0.25, cw*0.15, 0]
            t = Table(row_data, colWidths=col_w, rowHeights=[0.8*cm] + [None]*(len(row_data)-1))
            
            t_style = [
                ('BACKGROUND', (0,0), (3,0), colors.whitesmoke),
                ('VALIGN', (0,0), (3,-1), 'MIDDLE'),
                ('LINEABOVE', (0,0), (3,0), 1, colors.grey),
                ('LINEBELOW', (0,0), (3,0), 1, colors.grey),
                ('LINEBELOW', (0,-1), (3,-1), 1, colors.grey),
                ('LEFTPADDING', (4,0), (4,-1), 0),
                ('RIGHTPADDING', (4,0), (4,-1), 0),
            ]
            
            # Add dashed lines for body rows
            for i in range(1, len(row_data)):
                t_style.append(('LINEBELOW', (0,i), (3,i), 0.5, colors.grey, 1, (2, 2))) # dashed
                t_style.append(('LINEAFTER', (0,i), (2,i), 0.5, colors.lightgrey)) # vertical separators
                
            # Add vertical lines for header
            for i in range(0, 3):
                t_style.append(('LINEAFTER', (i,0), (i,0), 0.5, colors.lightgrey))
                
            t.setStyle(TableStyle(t_style))
            return t
        
        # Section 1
        story.append(Paragraph("■ 금월 진행업무", section_title))
        story.append(Spacer(1, 0.2*cm))
        
        t1 = make_section_table(
            data.get("this_month_tasks", []), 
            ["업무내용", "완결여부", "결과보고", "비 고"], 
            ["task", "status", "result", "note"], 5
        )
        story.append(t1)
        story.append(Spacer(1, 0.8*cm))
        
        # Section 2
        story.append(Paragraph("■ 차월 업무계획", section_title))
        story.append(Spacer(1, 0.2*cm))
        
        t2 = make_section_table(
            data.get("next_month_plans", []), 
            ["업무내용", "예정일자", "예상 문제점", "비 고"], 
            ["task", "date", "issue", "note"], 5
        )
        story.append(t2)
        story.append(Spacer(1, 0.8*cm))
        
        # Section 3
        story.append(Paragraph("■ 기타 의견", section_title))
        story.append(Spacer(1, 0.2*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey, spaceAfter=0.2*cm))
        story.append(Spacer(1, 0.2*cm))
        safe_opinions = data.get("opinions", "").replace("\n", "<br/>")
        story.append(Paragraph(safe_opinions, body_style))
        story.append(Spacer(1, 2.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceAfter=0.5*cm))
        
        doc.build(story)
        return pdf_path
    except Exception as e:
        print(f"PDF 생성 실패: {e}")
        return ""