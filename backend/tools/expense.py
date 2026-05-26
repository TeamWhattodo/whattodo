import os
import sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# (font_name, regular_path, bold_name, bold_path) 순서로 우선 탐색
_FONT_CANDIDATES: list[tuple[str, str, str, str]] = []
if sys.platform == "win32":
    _FONT_CANDIDATES = [
        ("Malgun", "C:/Windows/Fonts/malgun.ttf", "Malgun-Bold", "C:/Windows/Fonts/malgunbd.ttf"),
    ]
elif sys.platform == "darwin":
    _FONT_CANDIDATES = [
        ("AppleGothic", "/Library/Fonts/AppleGothic.ttf",
         "AppleGothic", "/Library/Fonts/AppleGothic.ttf"),
        ("AppleGothic", "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
         "AppleGothic", "/System/Library/Fonts/Supplemental/AppleGothic.ttf"),
    ]

_FONT_REGULAR = "Helvetica"
_FONT_BOLD    = "Helvetica-Bold"
_fonts_ready  = False


def _ensure_fonts() -> None:
    """처음 호출 시 한 번만 폰트를 등록한다."""
    global _FONT_REGULAR, _FONT_BOLD, _fonts_ready
    if _fonts_ready:
        return
    for reg_name, reg_path, bold_name, bold_path in _FONT_CANDIDATES:
        if os.path.exists(reg_path):
            pdfmetrics.registerFont(TTFont(reg_name, reg_path))
            if bold_name != reg_name and os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
            _FONT_REGULAR = reg_name
            _FONT_BOLD    = bold_name
            break
    _fonts_ready = True


OUTPUT_DIR    = "outputs"
TEMPLATE_PATH = "backend/db/data/Form/경비정산서 양식.xlsx"

CATEGORY_MAP = {
    "식비":   "식비",
    "숙박비": "숙박비",
    "유류비": "교통비",
    "출장비": "기타",
    "기타":   "기타",
}



def build_expense_report(items: list[dict], report_type: str = "출장비") -> dict:
    """
    ReceiptItem[] → 엑셀 + PDF 생성 후 ExpenseReport dict 반환
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    report_id = f"exp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    total     = sum(i["amount"] for i in items)

    xlsx_path = _write_xlsx(items, total, report_type, report_id)
    pdf_path  = _write_pdf(items, total, report_type, report_id)

    return {
        "id":           report_id,
        "created_at":   datetime.now().isoformat(),
        "report_type":  report_type,
        "items":        items,
        "total_amount": total,
        "xlsx_path":    xlsx_path,
        "pdf_path":     pdf_path,
    }


def _write_xlsx(items, total, report_type, report_id) -> str:
    path = f"{OUTPUT_DIR}/{report_id}.xlsx"

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 데이터 시작 행("내역")과 총액 행("경비 총액") 동적 탐지
    data_start_row = None
    total_row      = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "내역":
                data_start_row = cell.row
            elif cell.value == "경비 총액":
                total_row = cell.row

    for i, item in enumerate(items):
        row = data_start_row + i
        ws[f"C{row}"] = item["date"]
        ws[f"E{row}"] = CATEGORY_MAP.get(item["category"], "기타")
        ws[f"G{row}"] = item["amount"]
        ws[f"I{row}"] = item.get("memo") or ""

    ws[f"D{total_row}"] = total

    wb.save(path)
    return path


def _write_pdf(items, total, report_type, report_id) -> str:
    _ensure_fonts()

    path = f"{OUTPUT_DIR}/{report_id}.pdf"
    c    = canvas.Canvas(path, pagesize=A4)
    w, h = A4

    c.setFont(_FONT_BOLD, 16)
    c.drawString(50, h - 60, f"{report_type} 정산서")

    c.setFont(_FONT_REGULAR, 10)
    c.drawString(50, h - 90, f"작성일: {datetime.now().strftime('%Y-%m-%d')}")

    y = h - 130
    c.setFont(_FONT_BOLD, 10)
    c.drawString(50,  y, "날짜")
    c.drawString(130, y, "가맹점")
    c.drawString(280, y, "금액")
    c.drawString(360, y, "항목")

    y -= 20
    c.setFont(_FONT_REGULAR, 10)
    for item in items:
        c.drawString(50,  y, item["date"])
        c.drawString(130, y, item["merchant"])
        c.drawString(280, y, f"{item['amount']:,}원")
        c.drawString(360, y, item["category"])
        y -= 18

    y -= 10
    c.setFont(_FONT_BOLD, 10)
    c.drawString(130, y, "합계")
    c.drawString(280, y, f"{total:,}원")

    c.save()
    return path


if __name__ == "__main__":
    test_items = [
        {"date": "2026-05-12", "merchant": "GS칼텍스", "amount": 85000,
         "category": "유류비", "memo": "출장 이동"},
        {"date": "2026-05-12", "merchant": "롯데호텔", "amount": 42000,
         "category": "식비", "memo": "거래처 식사"},
    ]
    report = build_expense_report(test_items, "출장비")
    print(f"총액: {report['total_amount']:,}원")
    print(f"xlsx: {report['xlsx_path']}")
    print(f"pdf:  {report['pdf_path']}")
